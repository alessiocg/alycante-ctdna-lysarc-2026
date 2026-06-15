# -*- coding: utf-8 -*-
"""
V4 rebuild — Phase A : reconstruction de la verite-terrain MRD ctDNA depuis les
rapports phased-variant (FV), independamment du fichier Donnees_brutes2 historique.

Lit chaque rapport <NIP>_report.xlsx (onglet PV_Summary, 1 ligne/variant phase),
applique la logique d'exclusion par couleur de remplissage (orange = mauvaise qualite,
gris/beige = artefact expert, bleu = polymorphisme -> tout fill solide non-blanc exclu),
puis construit 4 versions de whitelist (WL) selon le critere de SELECTION (atteint sur
>=1 timepoint) :
    V1 : Common_UMI >= 1
    V2 : Common_UMI >= 10
    V3 : Common_UMI >= 1  & VAF_ratio > 0.3
    V4 : Common_UMI >= 10 & VAF_ratio > 0.3      <- reference (la plus propre)
MRD+ a un TP = >= 2 doublets de la WL detectes (Common_UMI >= 1) a ce TP.

Entrees (PHI, dans input/, git-ignorees) :
    input/ANALYSE_FV_OCTOBRE_2025/<NIP>_report.xlsx
Sortie (PHI, git-ignoree) :
    input/Donnees_brutes2_REBUILT_4WL.xlsx  (onglets REBUILT_4WL + WL_sizes)
"""
import os, sys, glob, re
import numpy as np
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _paths import input_path

FVDIR = input_path("ANALYSE_FV_OCTOBRE_2025")
OUT   = input_path("Donnees_brutes2_REBUILT_4WL.xlsx")

def num(v):
    try: return float(str(v).replace(',','.').replace('%',''))
    except: return None

def excluded(cell):
    """Exclu = tout remplissage solide non-blanc (lire .rgb ET .theme ET .indexed)."""
    f = cell.fill
    if not f or f.patternType != 'solid': return False           # pas de fill = blanc = GARDE
    c = f.fgColor
    if c.type == 'rgb' and isinstance(c.rgb, str): return c.rgb not in ('FFFFFFFF', '00000000')
    if c.type == 'theme': return not (c.theme == 0 and (c.tint or 0) == 0)
    if c.type == 'indexed': return c.indexed not in (64, 65)
    return True

def fukey(h):
    m = re.search(r'follow[\s_-]*up\s*(\d+)', h, re.I)
    if m: return 'FU' + m.group(1)
    if re.search(r'diagnos', h, re.I): return 'DIAG'
    return None

def norm_tp(raw):
    s = re.sub(r'follow[\s_-]*up\s*\d+', '', raw, flags=re.I)
    s = re.sub(r'diagnosis', '', s, flags=re.I).replace('_', ' ').strip().upper().rstrip('_').strip()
    s = re.sub(r'\bREF\b', '', s).strip()
    if s in ('DO', 'D O'): return 'D0'
    if s in ('Q14', '14'): return 'D14'
    if s in ('D-5', 'D5'): return 'D-5'
    if 'LEU' in s: return 'Leuca'
    if 'RECHUTE' in s or 'PROGRESS' in s or 'RELAPSE' in s: return 'Rechute'
    if 'GG' in s and 'INIT' in s: return 'Diag'
    if 'GG' in s: return 'Diag'
    m = re.match(r'(M\d+)', s)
    if m: return m.group(1)
    if s in ('D0', 'D14'): return s
    return s if s else 'Diag'

def parse(path):
    wb = openpyxl.load_workbook(path)
    if 'PV_Summary' not in wb.sheetnames: return None
    ws = wb['PV_Summary']; hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    umi = {}; vaf = {}; rat = {}; lab = {}
    for i, h in enumerate(hdr, 1):
        if not h: continue
        hs = str(h); k = fukey(hs)
        if k is None: continue
        if 'Common_UMI' in hs and 'Cumul' not in hs and 'Position' not in hs:
            umi[k] = i; lab[k] = norm_tp(hs.replace('Common_UMI', ''))
        elif 'VAF_ratio' in hs: rat[k] = i
        elif re.search(r'VAF(V1|V2)?$', hs) and 'ratio' not in hs.lower(): vaf.setdefault(k, i)
    keys = list(umi.keys())
    pvs = []
    for r in range(2, ws.max_row + 1):
        if all(ws.cell(row=r, column=c).value in (None, '') for c in range(1, 4)): continue
        pv = {'ex': excluded(ws.cell(row=r, column=1)), 'umi': {}, 'vaf': {}, 'rat': {}}
        for k in keys:
            pv['umi'][k] = num(ws.cell(row=r, column=umi[k]).value)
            pv['vaf'][k] = num(ws.cell(row=r, column=vaf[k]).value) if k in vaf else None
            pv['rat'][k] = num(ws.cell(row=r, column=rat[k]).value) if k in rat else None
        pvs.append(pv)
    return pvs, keys, lab

# Detection par-TP : chaque version applique SON critere
DET = {'V1': lambda pv, k: (pv['umi'].get(k) or 0) >= 1,
       'V2': lambda pv, k: (pv['umi'].get(k) or 0) >= 10,
       'V3': lambda pv, k: (pv['umi'].get(k) or 0) >= 1 and (pv['rat'].get(k) or 0) > 0.3,
       'V4': lambda pv, k: (pv['umi'].get(k) or 0) >= 10 and (pv['rat'].get(k) or 0) > 0.3}
rows = []; wlsize = []
files = [f for f in glob.glob(os.path.join(FVDIR, "*_report.xlsx"))
         if '_V2' not in os.path.basename(f) and not os.path.basename(f).startswith('~$')]
for f in sorted(files):
    nip = os.path.basename(f).split('_')[0]
    try:
        res = parse(f)
        if not res: continue
        pvs, keys, lab = res
        nonex = [pv for pv in pvs if not pv['ex']]
        wlsize.append(dict(NIP=nip, n_PV=len(pvs), n_exclus=sum(1 for pv in pvs if pv['ex']),
                           **{('WL_' + v): sum(1 for pv in nonex if any(R(pv, k) for k in keys)) for v, R in DET.items()}))
        for k in keys:
            tp = lab[k]
            rec = dict(NIP=nip, TP=tp, FU=k)
            for v, R in DET.items():
                d = sum(1 for pv in nonex if R(pv, k))
                rec['doublets_' + v] = d; rec['MRD_' + v] = 'POSITIF' if d >= 2 else 'NEGATIF'
            det = [pv for pv in nonex if (pv['umi'].get(k) or 0) >= 1]
            vafs = [pv['vaf'][k] for pv in det if pv['vaf'].get(k) is not None]
            rats = [pv['rat'][k] for pv in det if pv['rat'].get(k) is not None]
            rec['mean_VAF'] = round(np.mean(vafs), 4) if vafs else None
            rec['mean_VAF_ratio'] = round(np.mean(rats), 3) if rats else None
            rec['sum_common_umi'] = round(sum((pv['umi'].get(k) or 0) for pv in det), 1)
            rows.append(rec)
    except Exception as e:
        print(f"  {nip}: ERREUR {e}")

wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'REBUILT_4WL'
H = ['NIP', 'TP', 'FU'] + [f'{m}_{v}' for v in DET for m in ('doublets', 'MRD')] + ['mean_VAF', 'mean_VAF_ratio', 'sum_common_umi']
ws.append(H)
for r in rows: ws.append([r.get(h) for h in H])
ws2 = wb.create_sheet('WL_sizes'); H2 = ['NIP', 'n_PV', 'n_exclus', 'WL_V1', 'WL_V2', 'WL_V3', 'WL_V4']
ws2.append(H2)
for r in wlsize: ws2.append([r.get(h) for h in H2])
wb.save(OUT)
print(f"OK -> {OUT}\n{len(rows)} lignes (NIP x TP), {len(wlsize)} patients")
print("Taille WL totale (PV retenus, union des TP) :")
for v in DET: print(f"  {v}: {sum(w['WL_' + v] for w in wlsize)}")
print(f"  exclus (orange+gris/beige+bleu): {sum(w['n_exclus'] for w in wlsize)} / {sum(w['n_PV'] for w in wlsize)}")
for v in DET:
    pos = sum(1 for r in rows if r['MRD_' + v] == 'POSITIF')
    print(f"  TP MRD+ {v}: {pos}/{len(rows)}")
