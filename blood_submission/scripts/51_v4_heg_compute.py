# -*- coding: utf-8 -*-
"""
V4 rebuild — Phase B : calcul du marqueur heg (log10 hEG tumoral) par patient x TP,
pour les 4 versions de WL, en echelle PHYSIQUE (aucun offset, aucune convention) :

    MRD+ (>=2 doublets WL detectes) : heg = log10( fraction_poolee_UMI x cfDNA_hEG )
        fraction_poolee = somme(Common_UMI) / somme(Common_UMI / VAF) sur les doublets detectes
    MRD- (<2 doublets)              : heg = log10( cfDNA_hEG / (taille_WL_V x Profondeur_UMI) )
        = floor B = LOD 1/PCU_V, ou PCU_V = regions couvertes = taille WL (version) x profondeur

cfDNA_hEG = ADN_total_ng / 0.0033  (1 hEG = 3.3 pg). Source gold : onglet 'suivi cfDNA'.
La selection de la WL utilise le critere sur >=1 TP ; la detection par-TP = Common_UMI >= 1.

Entrees (PHI, input/, git-ignorees) :
    input/ANALYSE_FV_OCTOBRE_2025/<NIP>_report.xlsx
    input/reception_resultats_cfdna.xlsx   (onglet 'suivi cfDNA')
Sortie (PHI, git-ignoree) :
    input/data_lcmm_long_REBUILT_4WL.csv   (heg_V1..V4, mrd_V1..V4 par patient x TP)
"""
import os, sys, glob, re, math, csv, warnings
import numpy as np
import openpyxl
warnings.filterwarnings('ignore')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _paths import input_path

FVDIR = input_path("ANALYSE_FV_OCTOBRE_2025")
CFDNA = input_path("reception_resultats_cfdna.xlsx")
OUT   = input_path("data_lcmm_long_REBUILT_4WL.csv")

def num(v):
    try: return float(str(v).replace(',', '.').replace('%', ''))
    except: return None
def parse_vaf(v):   # "0.032 - 0.144" (VAFV1-VAFV2) -> moyenne ; "0.5" -> 0.5
    if v is None: return None
    nums = [float(x) for x in re.findall(r'\d+\.?\d*', str(v).replace(',', '.'))]
    return (sum(nums) / len(nums)) if nums else None
def nipn(s): return re.sub(r'[^A-Z0-9]', '', str(s).upper())
def nv(s):
    s = str(s).strip().lower().replace('_', ' ')
    if 'leuca' in s or 'leuka' in s: return 'Leuca'
    if 'relap' in s or 'rechut' in s or 'progress' in s: return 'Rechute'
    if 'gg' in s or 'diagnos' in s or 'init' in s: return 'Diag'
    s = s.replace(' ', '')
    if s.startswith('d-5'): return 'D-5'
    if s in ('d0', 'do') or s.startswith('d0'): return 'D0'
    if s in ('d14', 'q14', '14') or s.startswith('d14'): return 'D14'
    m = re.match(r'(m\d+)', s)
    if m: return m.group(1).upper()
    return s.upper()
TPTIME = {'Leuca': -1.5, 'D-5': -0.16, 'D0': 0.0, 'D14': 0.4599, 'M1': 1.02, 'M3': 2.99, 'M6': 6.03, 'M9': 9.05, 'M12': 11.99}

# cfDNA gold + NIP->SUBJID
wb = openpyxl.load_workbook(CFDNA, data_only=True); ws = wb['suivi cfDNA']
hd = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]; ci = {h: i for i, h in enumerate(hd, 1) if h}
cfdna = {}; nip2sub = {}
for r in range(2, ws.max_row + 1):
    nip = nipn(ws.cell(row=r, column=ci['Code labo']).value); sub = str(ws.cell(row=r, column=ci['randomisation']).value or '').replace('.0', '').strip()
    vis = nv(ws.cell(row=r, column=ci['Visite']).value); adn = num(ws.cell(row=r, column=ci['ADN total (ng)']).value)
    if nip and sub: nip2sub[nip] = sub
    if nip and adn and adn > 0: cfdna[(nip, vis)] = adn / 0.0033

def excluded(cell):
    f = cell.fill
    if not f or f.patternType != 'solid': return False
    c = f.fgColor
    if c.type == 'rgb' and isinstance(c.rgb, str): return c.rgb not in ('FFFFFFFF', '00000000')
    if c.type == 'theme': return not (c.theme == 0 and (c.tint or 0) == 0)
    if c.type == 'indexed': return c.indexed not in (64, 65)
    return True
def fukey(h):
    m = re.search(r'follow[\s_-]*up\s*(\d+)', h, re.I)
    if m: return 'FU' + m.group(1)
    if re.search(r'diagnos', h, re.I): return 'DIAG'
    return None
# SELECTION de la WL : critere atteint sur >=1 TP. Detection par-TP = UMI>=1 (presence).
SEL = {'V1': lambda u, r: (u or 0) >= 1, 'V2': lambda u, r: (u or 0) >= 10,
       'V3': lambda u, r: (u or 0) >= 1 and (r or 0) > 0.3,
       'V4': lambda u, r: (u or 0) >= 10 and (r or 0) > 0.3}

def qc_depth(wb):
    if 'QC' not in wb.sheetnames: return {}
    ws = wb['QC']; hd = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    ni = next((i for i, h in enumerate(hd, 1) if h and 'sample name' in str(h).lower()), 1)
    di = next((i for i, h in enumerate(hd, 1) if h and 'average umi depth' in str(h).lower()), None)
    d = {}
    if not di: return d
    for r in range(2, ws.max_row + 1):
        nm = str(ws.cell(row=r, column=ni).value or ''); dep = num(ws.cell(row=r, column=di).value)
        mm = re.search(r'(leuca|leuka|d-5|d14|d0|m\d+|gg ?rechute|rechute|progress|diag)', nm, re.I)
        if mm and dep: d[nv(mm.group(1))] = dep
    return d

out = []; cov = {'pos': 0, 'floor': 0, 'no_cfdna': 0}
files = [f for f in glob.glob(os.path.join(FVDIR, "*_report.xlsx")) if '_V2' not in os.path.basename(f) and not os.path.basename(f).startswith('~$')]
for f in sorted(files):
    nip = nipn(os.path.basename(f).split('_')[0]); sub = nip2sub.get(nip)
    if not sub: continue
    wb = openpyxl.load_workbook(f)
    if 'PV_Summary' not in wb.sheetnames: continue
    ws = wb['PV_Summary']; hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    umi = {}; vaf = {}; rat = {}; lab = {}
    for i, h in enumerate(hdr, 1):
        if not h: continue
        hs = str(h); k = fukey(hs)
        if k is None: continue
        if 'Common_UMI' in hs and 'Cumul' not in hs and 'Position' not in hs: umi[k] = i; lab[k] = nv(hs.replace('Common_UMI', ''))
        elif 'VAF_ratio' in hs: rat[k] = i
        elif re.search(r'VAF(V1|V2)?$', hs) and 'ratio' not in hs.lower(): vaf.setdefault(k, i)
    keys = list(umi.keys())
    pvs = []
    for r in range(2, ws.max_row + 1):
        if all(ws.cell(row=r, column=c).value in (None, '') for c in range(1, 4)): continue
        if excluded(ws.cell(row=r, column=1)): continue
        pv = {k: (num(ws.cell(row=r, column=umi[k]).value), parse_vaf(ws.cell(row=r, column=vaf[k]).value) if k in vaf else None, num(ws.cell(row=r, column=rat[k]).value) if k in rat else None) for k in keys}
        pvs.append(pv)
    depth = qc_depth(wb)
    # WL par version : PV (non exclu) atteignant le critere sur >=1 TP
    inWL = {v: [pv for pv in pvs if any(sel(pv[k][0], pv[k][2]) for k in keys)] for v, sel in SEL.items()}
    tpkeys = {}
    for k in keys:
        tp = lab[k]
        if tp in TPTIME: tpkeys.setdefault(tp, []).append(k)
    for tp, kk in tpkeys.items():
        cfh = cfdna.get((nip, tp))
        if not cfh: cov['no_cfdna'] += 1; continue
        row = {'randomisation': sub, 'NIP': nip, 'timepoint': tp, 'time': TPTIME[tp]}
        for v in SEL:
            det = [pv for pv in inWL[v] if any((pv[k][0] or 0) >= 1 for k in kk)]
            nd = len(det); num_umi = 0.0; den = 0.0
            for pv in det:
                for k in kk:
                    uu, vv, rr = pv[k]
                    if (uu or 0) >= 1 and vv and vv > 0: num_umi += uu; den += uu / vv
            frac = (num_umi / den) if den > 0 else 0.0
            if nd >= 2 and frac > 0:                       # MRD+ = >=2 doublets de la WL presents
                row['heg_' + v] = round(math.log10(frac * cfh), 4); row['mrd_' + v] = 1
            else:                                          # <2 = MRD- -> floor B (LOD = 1/PCU_V)
                dep = depth.get(tp) or (np.median([d for d in depth.values()]) if depth else 2000)
                pcu_v = max(len(inWL[v]), 1) * dep         # PCU_V = taille WL (version) x profondeur UMI = total couvert
                row['heg_' + v] = round(math.log10(cfh / pcu_v), 4); row['mrd_' + v] = 0
        out.append(row)
        cov['pos' if row['mrd_V1'] == 1 else 'floor'] += 1

H = ['randomisation', 'NIP', 'timepoint', 'time'] + [f'{m}_{v}' for v in SEL for m in ('heg', 'mrd')]
with open(OUT, 'w', encoding='utf-8', newline='') as fo:
    w = csv.DictWriter(fo, fieldnames=H); w.writeheader(); w.writerows(out)
print(f"OK -> {OUT}")
print(f"{len(out)} lignes (patient x TP), {len(set(r['randomisation'] for r in out))} patients")
print(f"TP avec signal V1: {cov['pos']} | floores V1: {cov['floor']} | TP sans cfDNA (sautes): {cov['no_cfdna']}")
for v in SEL:
    hh = [r['heg_' + v] for r in out]; pos = sum(1 for r in out if r['mrd_' + v] == 1)
    print(f"  {v}: heg q[0,50,100]=[{round(np.percentile(hh,0),2)},{round(np.percentile(hh,50),2)},{round(np.percentile(hh,100),2)}]  MRD+={pos}/{len(out)}")
